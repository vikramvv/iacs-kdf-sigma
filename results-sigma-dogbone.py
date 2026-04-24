import streamlit as st
import pandas as pd
import numpy as np
import re
import matplotlib.cm as cm
import matplotlib.colors as mcolors
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(layout="wide")
st.title("Dogbone Grid Statistics Viewer")
st.cache_data.clear()
uploaded_files = st.file_uploader(
    "Upload raw CSV files",
    type=["csv"],
    accept_multiple_files=True
)

if uploaded_files:
    sample_data = {}

    # --- Process each CSV ---
    for file in uploaded_files:
        df_raw = pd.read_csv(file)

        # Extract numeric values
        pattern = re.compile(r"RC\[current_value\]=([\d\.]+)")
        df_raw["NumericValue"] = df_raw["Value"].astype(str).str.extract(pattern).astype(float)
        df_raw = df_raw.dropna(subset=["NumericValue"]).copy()

        # Extract Row/Col
        if "Row" not in df_raw or "Col" not in df_raw:
            df_raw["Row"] = df_raw["Cell"].str.extract(r"R(\d+)").astype(int)
            df_raw["Col"] = df_raw["Cell"].str.extract(r"C(\d+)").astype(int)

        # Sample name
        sample_name = str(df_raw["Sample Name"].iloc[0]) if "Sample Name" in df_raw.columns else file.name

        # Stats
        stats = df_raw.groupby(["Row", "Col"])["NumericValue"].agg(["mean", "std"]).reset_index()
        sample_data[sample_name] = stats



    st.sidebar.header("Options")
    # --- Sidebar options ---
    unit_choice = st.sidebar.radio(
        "Units",
        ["IACS", "MS/m"],
        index=0  # default to IACS
    )
    # Determine scaling
    multiplier = 100/58 if unit_choice == "IACS" else 1.0
        # --- Colormap ---
    all_stats = pd.concat(sample_data.values(), ignore_index=True)
    # Scale all means before colormap normalization
    all_stats_scaled = all_stats.copy()
    all_stats_scaled["mean"] *= multiplier

    global_min = all_stats_scaled["mean"].min()
    global_max = all_stats_scaled["mean"].max()

    norm = mcolors.Normalize(vmin=global_min, vmax=global_max)
    cmap_choice = st.sidebar.selectbox(
        "Choose colormap",
        ["Yellow-Green (Excel Style)", "Viridis", "Plasma", "Coolwarm"]
    )
    if cmap_choice == "Yellow-Green (Excel Style)":
        cmap = mcolors.LinearSegmentedColormap.from_list("yellow_green", ["yellow", "green"])
    else:
        cmap = cm.get_cmap(cmap_choice.lower())
    norm = mcolors.Normalize(vmin=global_min, vmax=global_max)

    def get_color(val):
        r, g, b, _ = cmap(norm(val))
        return f"rgb({int(r*255)}, {int(g*255)}, {int(b*255)})", (r, g, b)

    # --- Build stacked table without R/C labels ---
    stacked_parts = []
    for sample_name, stats in sample_data.items():
        rows = sorted(stats["Row"].unique())
        cols = sorted(stats["Col"].unique())
        grid = np.full((len(rows), len(cols)), "", dtype=object)

        # Fill data
        for _, r in stats.iterrows():
            i = rows.index(r["Row"])
            j = cols.index(r["Col"])
            mean_scaled = r['mean'] * multiplier
            std_scaled = r['std'] * multiplier
            grid[i, j] = f"{mean_scaled:.2f} ± {std_scaled:.2f}"

        # DataFrame without row labels
        grid_df = pd.DataFrame(grid, columns=[f"C{c}" for c in cols])

        # Sample name row
        sample_row = pd.DataFrame([[f"Sample: {sample_name}"] + [""] * (len(cols)-1)],
                                  columns=grid_df.columns)
        # Spacer row
        spacer_row = pd.DataFrame([[""] * len(grid_df.columns)], columns=grid_df.columns)

        stacked_parts.extend([sample_row, grid_df, spacer_row])

    stacked_df = pd.concat(stacked_parts, ignore_index=True)

    # --- Coloring function ---
    def cell_color(val):
        if val == "" or pd.isna(val) or str(val).startswith("Sample:"):
            return ""
        try:
            mean_val = float(str(val).split()[0])
        except:
            return ""
        rgb, _ = get_color(mean_val)
        return f"background-color: {rgb};"

    # --- Display in Streamlit ---
    styled = stacked_df.style.applymap(cell_color)
    st.dataframe(styled, use_container_width=True)

    # --- Export to Excel ---
    wb = Workbook()
    ws = wb.active
    ws.append(list(stacked_df.columns))

    for r in stacked_df.itertuples(index=False):
        row_values = list(r)
        ws.append(row_values)
        for j, val in enumerate(row_values, start=1):
            if val == "" or str(val).startswith("Sample:"):
                continue
            try:
                mean_val = float(str(val).split()[0])
                rgb, (r_c, g_c, b_c) = get_color(mean_val)
                hex_color = f"{int(r_c*255):02X}{int(g_c*255):02X}{int(b_c*255):02X}"
                ws.cell(ws.max_row, j).fill = PatternFill(start_color=hex_color,
                                                          end_color=hex_color,
                                                          fill_type="solid")
            except:
                pass

    buf = BytesIO()
    wb.save(buf)
    st.download_button(
        "Download Excel with Colors",
        data=buf.getvalue(),
        file_name="dogbone_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("Upload one or more CSV files to begin.")
